import re
import secrets
import string
from datetime import timedelta, date
from io import BytesIO
import traceback

import openpyxl
from django.core.mail import send_mail
from django.db import transaction
from django.utils.timezone import is_aware

from .models import Remision

COLUMNAS_ESPERADAS = [
    'fecha', 'nombre', 'tipo_doc', 'doc', 'sexo', 'especialidad', 'edad', 'gest',
    'diagnostico', 'ta', 'fc', 'fr', 'tm', 'spo2', 'glasg', 'eps',
    'institucion_reporta', 'municipio', 'medico_refiere', 'medico_hudn',
    'radio_operador', 'observacion', 'aceptado', 'fecha_res',
]


def calcular_oportunidad(fecha, fecha_res) -> str:
    if fecha_res is None or fecha is None or fecha_res < fecha:
        return ''
    delta: timedelta = fecha_res - fecha
    total_minutos = int(delta.total_seconds() // 60)
    horas = total_minutos // 60
    minutos = total_minutos % 60
    return f'{horas:02d}:{minutos:02d}'


def es_registro_editable(remision) -> bool:
    if not remision.fecha:
        return False
    return (date.today() - remision.fecha.date()).days <= 1


def formatear_temperatura(valor: str) -> str:
    valor = valor.strip()
    if valor == 'NO REFIERE':
        return valor
    valor_limpio = valor.rstrip('°').strip()
    if re.match(r'^\d+(\.\d+)?$', valor_limpio):
        return f'{valor_limpio}°'
    return valor


def formatear_spo2(valor: str) -> str:
    valor = valor.strip()
    if valor == 'NO REFIERE':
        return valor
    valor_limpio = valor.rstrip('%').strip()
    if re.match(r'^\d+(\.\d+)?$', valor_limpio):
        return f'{valor_limpio}%'
    return valor


def validar_edad(valor: str) -> bool:
    if valor.strip().isdigit():
        valor = valor.strip() + " AÑOS"
    return bool(re.match(r'^\d+ (AÑOS|MESES|DIAS)$', valor.strip()))


def validar_ta(valor: str) -> bool:
    v = valor.strip()
    if v == 'NO REFIERE':
        return True
    return bool(re.match(r'^\d+/\d+$', v))


def validar_glasg(valor: str) -> bool:
    v = valor.strip()
    if v == 'NO REFIERE':
        return True
    return bool(re.match(r'^\d+/\d+$', v))


def obtener_remisiones(filtro: str, **kwargs):
    orden = kwargs.pop('orden', 'desc')
    order_field = '-fecha' if orden == 'desc' else 'fecha'
    qs = Remision.objects.all().order_by(order_field)

    if filtro == 'mes':
        mes = kwargs.get('mes')
        anio = kwargs.get('anio')
        qs = qs.filter(fecha__month=mes, fecha__year=anio)
    elif filtro == 'documento':
        doc = kwargs.get('doc', '').strip()
        if doc:
            from django.db.models import Q
            qs = qs.filter(Q(doc=doc) | Q(nombre__icontains=doc))

    return qs


def _excel_dt(dt):
    if dt is None:
        return None
    return dt.replace(tzinfo=None) if is_aware(dt) else dt


def exportar_a_excel(queryset) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Remisiones'
    ws.append(COLUMNAS_ESPERADAS + ['OPORTUNIDAD'])

    for r in queryset:
        oportunidad = calcular_oportunidad(r.fecha, r.fecha_res) if r.fecha_res else ''
        ws.append([
            _excel_dt(r.fecha), r.nombre, r.tipo_doc, r.doc, r.sexo,
            r.especialidad, r.edad, r.gest, r.diagnostico, r.ta, r.fc, r.fr, r.tm, r.spo2, r.glasg,
            r.eps, r.institucion_reporta, r.municipio, r.medico_refiere, r.medico_hudn,
            r.radio_operador, r.observacion, r.aceptado, _excel_dt(r.fecha_res), oportunidad,
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].number_format = 'DD/MM/YYYY HH:MM'
        row[23].number_format = 'DD/MM/YYYY HH:MM'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_password_temporal() -> str:
    alfabeto = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alfabeto) for _ in range(12))


def enviar_email_recuperacion(usuario, password_temp: str) -> None:
    asunto = 'Recuperación de contraseña — CRUE Remisiones'
    mensaje = (
        f'Hola {usuario.get_full_name() or usuario.username},\n\n'
        f'Tu contraseña temporal es: {password_temp}\n\n'
        'Por favor, inicia sesión y cámbiala de inmediato.\n\n'
        'Sistema CRUE Remisiones Pacientes'
    )
    send_mail(
        subject=asunto, message=mensaje, from_email=None,
        recipient_list=[usuario.email], fail_silently=False,
    )


def fechaHora(fecha_str, hora_str):
    from datetime import datetime as _datetime
    fecha_s = str(fecha_str or '').strip()
    hora_s = str(hora_str or '').strip()
    if not fecha_s or not hora_s:
        return None
    try:
        return _datetime.strptime(f'{fecha_s} {hora_s}', '%Y-%m-%d %H:%M')
    except ValueError:
        return None


def importar_desde_excel(archivo_o_path, usuario, sheet_name=None) -> dict:
    from . import utils_lg

    def sel(a, b):
        return a if a else b

    try:
        remFields = [f.name for f in Remision._meta.fields if f.name != 'id'][:-1]
        registros = []
        omitidos = 0
        df = utils_lg.excelToCsv(archivo_o_path, sheet_name=sheet_name)

        for r in df.itertuples(index=False):
            rowReg = [
                fechaHora(r[1], r[2]), *r[3:6], sel(r[6], r[7]), sel(r[8], r[9]), r[10],
                'OTRA', *r[11:25], sel(sel(r[25], r[26]), r[27]), fechaHora(r[28], r[29])
            ]
            fieldValueDic = dict(zip(remFields, rowReg))

            fecha = fieldValueDic.get('fecha')
            nombre = fieldValueDic.get('nombre')

            if Remision.objects.filter(fecha=fecha, nombre=nombre).exists():
                omitidos += 1
                continue

            registros.append(Remision(**fieldValueDic))

        with transaction.atomic():
            Remision.objects.bulk_create(registros)

        return {'ok': True, 'importados': len(registros), 'omitidos': omitidos}

    except Exception as ex:
        traceback.print_exc()
        return {'ok': False, 'error': str(ex)}
