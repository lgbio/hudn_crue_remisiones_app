"""
Lógica de negocio desacoplada de las vistas.

Módulo services.py — CRUE Remisiones Pacientes
"""

import re
import secrets
import string

from datetime import timedelta
from datetime import date

from io import BytesIO
import traceback

import openpyxl
from django.core.mail import send_mail
from django.db import transaction
from django.utils import timezone

from .models import Remision

# ---------------------------------------------------------------------------
# Columnas esperadas en el archivo Excel de importación/exportación
# ---------------------------------------------------------------------------

COLUMNAS_ESPERADAS = [
	'fecha', 'nombre', 'tipo_doc', 'doc', 'sexo', 'especialidad', 'edad', 'gest',
	'diagnostico', 'ta', 'fc', 'fr', 'tm', 'spo2', 'glasg', 'eps',
	'institucion_reporta', 'municipio', 'medico_refiere', 'medico_hudn',
	'radio_operador', 'observacion', 'aceptado', 'fecha_res',
]


# ---------------------------------------------------------------------------
# 3.1 — Cálculo de oportunidad y editabilidad
# ---------------------------------------------------------------------------

def calcular_oportunidad(fecha, fecha_res) -> str:
	"""
	Calcula la diferencia entre fecha_res y fecha en formato HH:MM.

	Retorna cadena vacía si fecha_res es None o fecha_res < fecha.
	"""
	if fecha_res is None or fecha_res < fecha:
		return ''
	delta: timedelta = fecha_res - fecha
	total_minutos = int(delta.total_seconds() // 60)
	horas = total_minutos // 60
	minutos = total_minutos % 60
	return f'{horas:02d}:{minutos:02d}'


def es_registro_editable(remision) -> bool:
	"""
	Retorna True si el registro tiene como máximo 1 día de antigüedad.
	Es decir, es editable si fecha.date() es hoy o ayer.
	"""
	if not remision.fecha:
		return False

	dias_diferencia = (date.today() - remision.fecha.date()).days
	return dias_diferencia <= 1

#def es_registro_editable(remision) -> bool:
#	"""
#	Retorna True si la fecha del registro corresponde al día de hoy en la
#	zona horaria local configurada en Django (America/Bogota).
#
#	Los registros históricos (fecha.date() < hoy) son de solo lectura.
#	"""
#	fecha_local = remision.fecha.astimezone(
#		timezone.get_current_timezone()
#	).date()
#	return fecha_local == timezone.localdate()


# ---------------------------------------------------------------------------
# 3.2 — Autoformato de signos vitales
# ---------------------------------------------------------------------------

def formatear_temperatura(valor: str) -> str:
	"""
	Agrega el símbolo de grados (°) si el valor es numérico y no lo tiene ya.
	Idempotente: no duplica el símbolo.
	Pasa 'NO REFIERE' sin cambios.
	"""
	valor = valor.strip()
	if valor == 'NO REFIERE':
		return valor
	# Quitar símbolo existente para garantizar idempotencia
	valor_limpio = valor.rstrip('\u00b0').strip()
	if re.match(r'^\d+(\.\d+)?$', valor_limpio):
		return f'{valor_limpio}\u00b0'
	return valor


def formatear_spo2(valor: str) -> str:
	"""
	Agrega el símbolo de porcentaje (%) si el valor es numérico y no lo tiene ya.
	Idempotente: no duplica el símbolo.
	Pasa 'NO REFIERE' sin cambios.
	"""
	valor = valor.strip()
	if valor == 'NO REFIERE':
		return valor
	valor_limpio = valor.rstrip('%').strip()
	if re.match(r'^\d+(\.\d+)?$', valor_limpio):
		return f'{valor_limpio}%'
	return valor


# ---------------------------------------------------------------------------
# 3.3 — Validación de campos
# ---------------------------------------------------------------------------

def validar_edad(valor: str) -> bool:
	"""Valida formato: número seguido de 'AÑOS', 'MESES' o 'DIAS'."""
	if valor.strip().isdigit():
		valor = valor.strip() + " AÑOS"

	return bool(re.match(r'^\d+ (AÑOS|MESES|DIAS)$', valor.strip()))


def validar_ta(valor: str) -> bool:
	"""Valida formato fracción (ej. 120/80) o 'NO REFIERE'."""
	v = valor.strip()
	if v == 'NO REFIERE':
		return True
	return bool(re.match(r'^\d+/\d+$', v))


def validar_glasg(valor: str) -> bool:
	"""Valida formato fracción (ej. 15/15) o 'NO REFIERE'."""
	v = valor.strip()
	if v == 'NO REFIERE':
		return True
	return bool(re.match(r'^\d+/\d+$', v))


def validar_signo_vital_numerico(valor: str) -> bool:
	"""Valida que el valor sea numérico o 'NO REFIERE' (para FC, FR)."""
	v = valor.strip()
	if v == 'NO REFIERE':
		return True
	return bool(re.match(r'^\d+(\.\d+)?$', v))


# ---------------------------------------------------------------------------
# 3.4 — Filtros de consulta
# ---------------------------------------------------------------------------

def obtener_remisiones(filtro: str, **kwargs):
	"""
	Aplica el filtro activo y retorna un QuerySet ordenado por fecha.

	filtro='mes':		kwargs = {mes: int, anio: int}
	filtro='documento': kwargs = {doc: str} — searches by doc number OR patient name

	orden: 'desc' (default, newest first) or 'asc' (oldest first)
	"""
	orden = kwargs.pop('orden', 'desc')
	order_field = '-fecha' if orden == 'desc' else 'fecha'
	qs = Remision.objects.all().order_by(order_field)

	if filtro == 'mes':
		mes = kwargs.get('mes')
		anio = kwargs.get('anio')
		qs = qs.filter(fecha__month=mes, fecha__year=anio)

	elif filtro == 'documento':
		doc = kwargs.get('doc', '').strip()
		if doc:
			# Search by document number OR patient name (case-insensitive)
			from django.db.models import Q
			qs = qs.filter(Q(doc=doc) | Q(nombre__icontains=doc))

	return qs



# ---------------------------------------------------------------------------
# 3.6 — Exportación a Excel
# ---------------------------------------------------------------------------
from io import BytesIO
from django.utils.timezone import is_aware
import openpyxl


def _excel_dt(dt):
	"""
	Convierte datetimes timezone-aware en datetimes compatibles con Excel.
	Excel no soporta tzinfo.
	"""
	if dt is None:
		return None

	return dt.replace(tzinfo=None) if is_aware(dt) else dt


def exportar_a_excel(queryset) -> BytesIO:
	"""
	Genera un archivo Excel con todos los campos de Remision,
	incluyendo la columna OPORTUNIDAD calculada dinámicamente.
	"""
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = 'Remisiones'

	encabezados = COLUMNAS_ESPERADAS + ['OPORTUNIDAD']
	ws.append(encabezados)

	for r in queryset:
		oportunidad = calcular_oportunidad(r.fecha, r.fecha_res) if r.fecha_res else ''

		ws.append([
			_excel_dt(r.fecha), r.nombre, r.tipo_doc, r.doc, r.sexo,
			r.especialidad, r.edad, r.gest, r.diagnostico, r.ta, r.fc, r.fr, r.tm, r.spo2, r.glasg,
			r.eps, r.institucion_reporta, r.municipio, r.medico_refiere, r.medico_hudn,
			r.radio_operador, r.observacion, r.aceptado, _excel_dt(r.fecha_res),
			oportunidad,
		])

	# Formato visual para columnas fecha y fecha_res
	for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
		row[0].number_format = 'DD/MM/YYYY HH:MM'	# fecha
		row[23].number_format = 'DD/MM/YYYY HH:MM'	# fecha_res

	buffer = BytesIO()
	wb.save(buffer)
	buffer.seek(0)

	return buffer


# ---------------------------------------------------------------------------
# 3.7 — Utilidades de usuario
# ---------------------------------------------------------------------------

def generar_password_temporal() -> str:
	"""
	Genera una contraseña temporal de 12 caracteres usando letras y dígitos.
	Usa el módulo `secrets` para garantizar aleatoriedad criptográficamente segura.
	"""
	alfabeto = string.ascii_letters + string.digits
	return ''.join(secrets.choice(alfabeto) for _ in range(12))


def enviar_email_recuperacion(usuario, password_temp: str) -> None:
	"""
	Envía un correo electrónico al usuario con su contraseña temporal.
	Usa el backend de email configurado en Django (SMTP en producción,
	locmem en pruebas).
	"""
	asunto = 'Recuperación de contraseña — CRUE Remisiones'
	mensaje = (
		f'Hola {usuario.get_full_name() or usuario.username},\n\n'
		f'Tu contraseña temporal es: {password_temp}\n\n'
		'Por favor, inicia sesión y cámbiala de inmediato.\n\n'
		'Sistema CRUE Remisiones Pacientes'
	)
	send_mail(
		subject=asunto,
		message=mensaje,
		from_email=None,  # Usa DEFAULT_FROM_EMAIL de settings
		recipient_list=[usuario.email],
		fail_silently=False,
	)

# ---------------------------------------------------------------------------
# For truncating text larger than field max_length
# ---------------------------------------------------------------------------
# Build dict: {field_name: max_length}
fieldMaxLens = {
	f.name: f.max_length
	for f in Remision._meta.fields
	if hasattr(f, 'max_length') and f.max_length
}

def truncate_fields(data: dict) -> dict:
	for field, max_len in fieldMaxLens.items():
		val = data.get(field)

		if isinstance(val, str) and len(val) > max_len:
			print(f"+++ Truncando campo '{field}' "
				  f"({len(val)} -> {max_len})")

			data[field] = val[:max_len]

	return data

## ---------------------------------------------------------------------------
## 3.5 — Importación desde Excel
## ---------------------------------------------------------------------------
def importar_desde_excel (archivo_o_path, usuario, sheet_name=None) -> dict:
	from . import utils_lg
	import tempfile, os
	#-- Select a or b
	def sel(a, b):
		return a if a else b
	try:
		# If archivo_o_path is a file-like object, save to temp file
		temp_path = None
		if hasattr(archivo_o_path, 'read'):
			with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
				tmp.write(archivo_o_path.read())
				temp_path = tmp.name
			excel_path = temp_path
		else:
			excel_path = str(archivo_o_path)

		# Create registros from dataframe
		remFields = [f.name for f in Remision._meta.fields if f.name != "id"][:-1]
		registros = []
		omitidos  = 0
		df		  = utils_lg.excelToCsv(archivo_o_path, sheet_name=sheet_name)

		for r in df.itertuples(index=False):
			rowReg = [
				fechaHora(r[1], r[2]), *r[3:6], sel(r[6], r[7]), sel(r[8], r[9]), r[10],
				"OTRA", *r[11:25], sel(sel(r[25], r[26]), r[27]), fechaHora(r[28], r[29])
			]
			fieldValueDic = dict(zip(remFields, rowReg))

			# Truncate oversized text fields
			fieldValueDic = truncate_fields(fieldValueDic)
			print(f"\n+++ {fieldValueDic=}")

			fecha  = fieldValueDic.get('fecha')
			nombre = fieldValueDic.get('nombre')

			# Skip if a record with same fecha+nombre already exists
			if Remision.objects.filter (fecha=fecha, nombre=nombre).exists():
				print (f"+++ Omitido (ya existe): {fecha=}, {nombre=}")
				omitidos += 1
				continue

			registros.append(Remision(**fieldValueDic))

		# Bulk insert only new registros
		with transaction.atomic():
			Remision.objects.bulk_create(registros)

		return {
			'ok'		: True,
			'importados': len(registros),
			'omitidos'	: omitidos,
		}

	except Exception as ex:
		print(f"+++ Error importando desde excel  {ex=}")
		traceback.print_exc()
		return {'ok': False, 'error': str(ex)}

#def importar_desde_excel_v2 (archivo_o_path, usuario, sheet_name=None) -> dict:
#	from . import utils_lg
#	import tempfile, os
#
#	#-- Select a or b
#	def sel (a, b):
#		return a if a else b
#
#	try:
#		# If archivo_o_path is a file-like object, save to temp file
#		temp_path = None
#		if hasattr (archivo_o_path, 'read'):
#			with tempfile.NamedTemporaryFile (suffix='.xlsx', delete=False) as tmp:
#				tmp.write (archivo_o_path.read())
#				temp_path = tmp.name
#			excel_path = temp_path
#		else:
#			excel_path = str(archivo_o_path)
#
#		# Create registros from dataframe
#		remFields = [f.name for f in Remision._meta.fields if f.name != "id"][:-1] # Remision field names
#		registros = []	 # Regs for Remision
#		df		  = utils_lg.excelToCsv (archivo_o_path, sheet_name=sheet_name)
#		for r in df.itertuples (index=False):
#			rowReg = [
#				fechaHora (r[1],r[2]), *r[3:6], sel(r[6],r[7]), sel(r[8],r[9]), r[10], 
#				"OTRA", *r[11:25], sel(sel(r[25],r[26]),r[27]), fechaHora (r[28],r[29])
#			]
#			fieldValueDic = dict(zip(remFields, rowReg))
#			print (f"\n+++ {fieldValueDic=}")
#			remReg = Remision (**fieldValueDic)
#			registros.append (remReg)
#			
#		# Bulk update of 'registros' to Remisions
#		with transaction.atomic():
#			Remision.objects.bulk_create (registros)
#		
#		return {'ok': True, 'importados': len(registros)}		
#	except Exception as ex:
#		print (f"+++ Error importando desde excel  {ex=}")
#		traceback.print_exc()
#		return {'ok': False, 'error': str(ex)}		

## ---------------------------------------------------------------------------
## Helpers para importar_desde_excel — mapeo de campos exclusivos
## ---------------------------------------------------------------------------
def fechaHora (fecha_str, hora_str):
	"""
	Combines date string 'YYYY-MM-DD' and time string 'HH:MM' into a datetime.
	Returns None if either is empty/None.
	Returns None if combination is invalid.
	"""
	from datetime import datetime as _datetime
	fecha_s = str(fecha_str or '').strip()
	hora_s = str(hora_str or '').strip()
	if not fecha_s or not hora_s:
		return None
	try:
		return _datetime.strptime(f'{fecha_s} {hora_s}', '%Y-%m-%d %H:%M')
	except ValueError:
		return None


