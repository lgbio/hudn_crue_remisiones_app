"""
Lógica de negocio desacoplada de las vistas.

Módulo services.py — CRUE Remisiones Pacientes
"""

import re
import secrets
import string
from datetime import timedelta
from io import BytesIO

import openpyxl
from django.core.mail import send_mail
from django.db import transaction
from django.utils import timezone

from .models import Remision

# ---------------------------------------------------------------------------
# Columnas esperadas en el archivo Excel de importación/exportación
# ---------------------------------------------------------------------------

COLUMNAS_ESPERADAS = [
    'fecha', 'nombre', 'tipo_doc', 'doc', 'sexo', 'especialidad', 'edad', 'gest',
    'diagnostico', 'ta', 'fc', 'fr', 'tm', 'spo2', 'glasg', 'eps',
    'institucion_reporta', 'municipio', 'medico_refiere', 'medico_hudn',
    'radio_operador', 'observacion', 'aceptado', 'fecha_res',
]


# ---------------------------------------------------------------------------
# 3.1 — Cálculo de oportunidad y editabilidad
# ---------------------------------------------------------------------------

def calcular_oportunidad(fecha, fecha_res) -> str:
    """
    Calcula la diferencia entre fecha_res y fecha en formato HH:MM.

    Retorna cadena vacía si fecha_res es None o fecha_res < fecha.
    """
    if fecha_res is None or fecha_res < fecha:
        return ''
    delta: timedelta = fecha_res - fecha
    total_minutos = int(delta.total_seconds() // 60)
    horas = total_minutos // 60
    minutos = total_minutos % 60
    return f'{horas:02d}:{minutos:02d}'


def es_registro_editable(remision) -> bool:
    """
    Retorna True si la fecha del registro corresponde al día de hoy en la
    zona horaria local configurada en Django (America/Bogota).

    Los registros históricos (fecha.date() < hoy) son de solo lectura.
    """
    fecha_local = remision.fecha.astimezone(
        timezone.get_current_timezone()
    ).date()
    return fecha_local == timezone.localdate()


# ---------------------------------------------------------------------------
# 3.2 — Autoformato de signos vitales
# ---------------------------------------------------------------------------

def formatear_temperatura(valor: str) -> str:
    """
    Agrega el símbolo de grados (°) si el valor es numérico y no lo tiene ya.
    Idempotente: no duplica el símbolo.
    Pasa 'NO REFIERE' sin cambios.
    """
    valor = valor.strip()
    if valor == 'NO REFIERE':
        return valor
    # Quitar símbolo existente para garantizar idempotencia
    valor_limpio = valor.rstrip('\u00b0').strip()
    if re.match(r'^\d+(\.\d+)?$', valor_limpio):
        return f'{valor_limpio}\u00b0'
    return valor


def formatear_spo2(valor: str) -> str:
    """
    Agrega el símbolo de porcentaje (%) si el valor es numérico y no lo tiene ya.
    Idempotente: no duplica el símbolo.
    Pasa 'NO REFIERE' sin cambios.
    """
    valor = valor.strip()
    if valor == 'NO REFIERE':
        return valor
    valor_limpio = valor.rstrip('%').strip()
    if re.match(r'^\d+(\.\d+)?$', valor_limpio):
        return f'{valor_limpio}%'
    return valor


# ---------------------------------------------------------------------------
# 3.3 — Validación de campos
# ---------------------------------------------------------------------------

def validar_edad(valor: str) -> bool:
    """Valida formato: número seguido de 'AÑOS', 'MESES' o 'DIAS'."""
    if valor.strip().isdigit():
        valor = valor.strip() + " AÑOS"

    return bool(re.match(r'^\d+ (AÑOS|MESES|DIAS)$', valor.strip()))


def validar_ta(valor: str) -> bool:
    """Valida formato fracción (ej. 120/80) o 'NO REFIERE'."""
    v = valor.strip()
    if v == 'NO REFIERE':
        return True
    return bool(re.match(r'^\d+/\d+$', v))


def validar_glasg(valor: str) -> bool:
    """Valida formato fracción (ej. 15/15) o 'NO REFIERE'."""
    v = valor.strip()
    if v == 'NO REFIERE':
        return True
    return bool(re.match(r'^\d+/\d+$', v))


def validar_signo_vital_numerico(valor: str) -> bool:
    """Valida que el valor sea numérico o 'NO REFIERE' (para FC, FR)."""
    v = valor.strip()
    if v == 'NO REFIERE':
        return True
    return bool(re.match(r'^\d+(\.\d+)?$', v))


# ---------------------------------------------------------------------------
# 3.4 — Filtros de consulta
# ---------------------------------------------------------------------------

def obtener_remisiones(filtro: str, **kwargs):
    """
    Aplica el filtro activo y retorna un QuerySet ordenado por fecha ascendente.

    filtro='mes':       kwargs = {mes: int, anio: int}
    filtro='rango':     kwargs = {desde: date, hasta: date}
    filtro='documento': kwargs = {doc: str}
    """
    qs = Remision.objects.all().order_by('fecha')

    if filtro == 'mes':
        mes = kwargs.get('mes')
        anio = kwargs.get('anio')
        qs = qs.filter(fecha__month=mes, fecha__year=anio)

    elif filtro == 'rango':
        desde = kwargs.get('desde')
        hasta = kwargs.get('hasta')
        qs = qs.filter(fecha__date__gte=desde, fecha__date__lte=hasta)

    elif filtro == 'documento':
        doc = kwargs.get('doc', '').strip()
        qs = qs.filter(doc=doc)

    return qs


# ---------------------------------------------------------------------------
# 3.5 — Importación desde Excel
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Helpers para importar_desde_excel_v2 — mapeo de campos exclusivos
# ---------------------------------------------------------------------------

def _mapear_exclusivo_sexo(row) -> str:
    """
    Maps exclusive SEXO columns from DataFrame row.
    sexo_m='SI' → 'M'; sexo_f='SI' → 'F'
    Raises ValueError if both or neither have 'SI'.
    """
    sexo_m = str(row.get('sexo_m', '') or '').strip().upper()
    sexo_f = str(row.get('sexo_f', '') or '').strip().upper()
    si_m = sexo_m == 'SI'
    si_f = sexo_f == 'SI'
    if si_m and si_f:
        raise ValueError("SEXO: ambos sexo_m y sexo_f tienen 'SI' (exclusividad violada)")
    if not si_m and not si_f:
        raise ValueError("SEXO: ninguno de sexo_m o sexo_f tiene 'SI' (campo requerido)")
    return 'M' if si_m else 'F'


def _mapear_exclusivo_gest(row) -> str:
    """
    Maps exclusive GEST columns from DataFrame row.
    gestante_si='SI' → 'SI'; gestante_no='SI' → 'NO'
    Raises ValueError if both or neither have 'SI'.
    """
    gest_si = str(row.get('gestante_si', '') or '').strip().upper()
    gest_no = str(row.get('gestante_no', '') or '').strip().upper()
    si_si = gest_si == 'SI'
    si_no = gest_no == 'SI'
    if si_si and si_no:
        raise ValueError("GEST: ambos gestante_si y gestante_no tienen 'SI' (exclusividad violada)")
    if not si_si and not si_no:
        raise ValueError("GEST: ninguno de gestante_si o gestante_no tiene 'SI' (campo requerido)")
    return 'SI' if si_si else 'NO'


def _mapear_exclusivo_aceptado(row) -> str:
    """
    Maps exclusive ACEPTADO columns from DataFrame row.
    aceptado_si='SI' → 'SI'; aceptado_no='SI' → 'NO'; aceptado_urg_vital='SI' → 'URG VITAL'
    Raises ValueError if more than one or none have 'SI'.
    """
    a_si = str(row.get('aceptado_si', '') or '').strip().upper()
    a_no = str(row.get('aceptado_no', '') or '').strip().upper()
    a_uv = str(row.get('aceptado_urg_vital', '') or '').strip().upper()
    count = sum([a_si == 'SI', a_no == 'SI', a_uv == 'SI'])
    if count > 1:
        raise ValueError("ACEPTADO: más de un campo tiene 'SI' (exclusividad violada)")
    if count == 0:
        raise ValueError("ACEPTADO: ningún campo tiene 'SI' (campo requerido)")
    if a_si == 'SI':
        return 'SI'
    if a_no == 'SI':
        return 'NO'
    return 'URG VITAL'


def _combinar_fecha_hora(fecha_str, hora_str):
    """
    Combines date string 'YYYY-MM-DD' and time string 'HH:MM' into a datetime.
    Returns None if either is empty/None.
    Returns None if combination is invalid.
    """
    from datetime import datetime as _datetime
    fecha_s = str(fecha_str or '').strip()
    hora_s = str(hora_str or '').strip()
    if not fecha_s or not hora_s:
        return None
    try:
        return _datetime.strptime(f'{fecha_s} {hora_s}', '%Y-%m-%d %H:%M')
    except ValueError:
        return None


def importar_desde_excel_v2(archivo_o_path, usuario) -> dict:
    """
    Importa registros desde un archivo Excel FRALG-062 usando excelToDataframe.

    Retorna {'ok': True, 'importados': N} o {'ok': False, 'error': mensaje}.
    """
    import tempfile
    import os
    from .utils import excelToDataframe, SheetStructureError

    # If archivo_o_path is a file-like object, save to temp file
    temp_path = None
    try:
        if hasattr(archivo_o_path, 'read'):
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp.write(archivo_o_path.read())
                temp_path = tmp.name
            excel_path = temp_path
        else:
            excel_path = str(archivo_o_path)

        # 1. Validate structure and get DataFrame
        try:
            df = excelToDataframe(excel_path)
        except SheetStructureError as e:
            return {'ok': False, 'error': str(e)}
        except FileNotFoundError as e:
            return {'ok': False, 'error': str(e)}
        except Exception as e:
            return {'ok': False, 'error': f'Error al leer el archivo: {e}'}

        # 2. Map and validate each row
        registros = []
        for idx, row in df.iterrows():
            fila_num = idx + 1  # 1-based for user messages
            row_dict = row.to_dict()

            # Map exclusive fields
            try:
                sexo = _mapear_exclusivo_sexo(row_dict)
            except ValueError as e:
                return {'ok': False, 'error': f'Fila {fila_num}: {e}'}

            try:
                gest = _mapear_exclusivo_gest(row_dict)
            except ValueError as e:
                return {'ok': False, 'error': f'Fila {fila_num}: {e}'}

            try:
                aceptado = _mapear_exclusivo_aceptado(row_dict)
            except ValueError as e:
                return {'ok': False, 'error': f'Fila {fila_num}: {e}'}

            # Combine fecha (fecha + hora)
            fecha_str = str(row_dict.get('fecha', '') or '').strip()
            hora_str = str(row_dict.get('hora', '') or '').strip()
            fecha_val = _combinar_fecha_hora(fecha_str, hora_str)
            if fecha_val is None:
                return {'ok': False, 'error': f'Fila {fila_num}: FECHA inválida (fecha={fecha_str!r}, hora={hora_str!r})'}

            # Combine fecha_res (aceptado_fecha + aceptado_hora) — NULL if either empty
            aceptado_fecha = str(row_dict.get('aceptado_fecha', '') or '').strip()
            aceptado_hora = str(row_dict.get('aceptado_hora', '') or '').strip()
            fecha_res_val = _combinar_fecha_hora(aceptado_fecha, aceptado_hora)
            # fecha_res can be None — that's valid per Req. 15.11

            # Validate format fields
            edad_val = str(row_dict.get('edad', '') or '').strip()
            if edad_val and not validar_edad(edad_val):
                return {'ok': False, 'error': f'Fila {fila_num}: EDAD formato inválido ({edad_val!r})'}

            ta_val = str(row_dict.get('sv_ta', '') or '').strip()
            if ta_val and not validar_ta(ta_val):
                return {'ok': False, 'error': f'Fila {fila_num}: TA formato inválido ({ta_val!r})'}

            glasg_val = str(row_dict.get('sv_glasgow', '') or '').strip()
            if glasg_val and not validar_glasg(glasg_val):
                return {'ok': False, 'error': f'Fila {fila_num}: GLASG formato inválido ({glasg_val!r})'}

            # Check for duplicates
            doc_val = str(row_dict.get('identificacion', '') or '').strip()
            if Remision.objects.filter(doc=doc_val, fecha=fecha_val).exists():
                return {'ok': False, 'error': f'Fila {fila_num}: Duplicado encontrado (DOC={doc_val!r}, FECHA={fecha_val})'}

            def s(key):
                v = row_dict.get(key)
                if v is None or (hasattr(v, '__class__') and v.__class__.__name__ == 'NAType'):
                    return ''
                return str(v).strip()

            registros.append(Remision(
                fecha=fecha_val,
                nombre=s('nombre_paciente'),
                tipo_doc=s('tipo_documento') or 'CC',
                doc=doc_val,
                sexo=sexo,
                edad=edad_val,
                gest=gest,
                diagnostico=s('diagnostico'),
                ta=ta_val,
                fc=s('sv_fc'),
                fr=s('sv_fr'),
                tm=s('sv_temperatura'),
                spo2=s('sv_spo2'),
                glasg=glasg_val or '15/15',
                eps=s('eps'),
                institucion_reporta=s('institucion_reporta'),
                municipio=s('municipio'),
                medico_refiere=s('medico_refiere'),
                medico_hudn=s('medico_hudn_confirma'),
                radio_operador=s('radio_operador'),
                observacion=s('observacion'),
                aceptado=aceptado,
                fecha_res=fecha_res_val,
                created_by=usuario,
            ))

        # 3. Atomic bulk create
        with transaction.atomic():
            Remision.objects.bulk_create(registros)

        return {'ok': True, 'importados': len(registros)}

    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


def importar_desde_excel(archivo, usuario) -> dict:
    """
    Importa registros desde un archivo Excel de forma atómica.

    Retorna {'ok': True, 'importados': N} o {'ok': False, 'error': mensaje}.
    Solo lee la primera hoja del libro.
    """
    try:
        wb = openpyxl.load_workbook(archivo)
    except Exception as exc:
        return {'ok': False, 'error': f'No se pudo abrir el archivo: {exc}'}

    ws = wb.active  # Primera hoja

    # 1. Validar encabezados
    encabezados = [cell.value for cell in ws[1]]
    if encabezados != COLUMNAS_ESPERADAS:
        faltantes = set(COLUMNAS_ESPERADAS) - set(encabezados)
        return {
            'ok': False,
            'error': f'Encabezados incorrectos. Faltan: {faltantes}',
        }

    filas = list(ws.iter_rows(min_row=2, values_only=True))

    # 2. Validar duplicados (doc + fecha ya existentes en BD)
    duplicados = []
    for i, fila in enumerate(filas, start=2):
        doc_val = str(fila[3]) if fila[3] is not None else ''
        fecha_val = fila[0]
        if fecha_val and Remision.objects.filter(doc=doc_val, fecha=fecha_val).exists():
            duplicados.append(f'Fila {i}: DOC={doc_val}, FECHA={fecha_val}')
    if duplicados:
        return {'ok': False, 'error': f'Duplicados encontrados: {duplicados}'}

    # 3. Validar formatos de cada fila
    errores_formato = []
    for i, fila in enumerate(filas, start=2):
        edad_val = str(fila[5]) if fila[5] is not None else ''
        if edad_val and not validar_edad(edad_val):
            errores_formato.append(f'Fila {i}: EDAD formato inválido ({edad_val})')
        ta_val = str(fila[8]) if fila[8] is not None else ''
        if ta_val and not validar_ta(ta_val):
            errores_formato.append(f'Fila {i}: TA formato inválido ({ta_val})')
    if errores_formato:
        return {'ok': False, 'error': f'Errores de formato: {errores_formato}'}

    # 4. Importación atómica
    with transaction.atomic():
        for fila in filas:
            Remision.objects.create(
                fecha=fila[0],
                nombre=fila[1],
                tipo_doc=fila[2],
                doc=str(fila[3]),
                sexo=fila[4],
                edad=fila[5],
                gest=fila[6],
                diagnostico=fila[7] or '',
                ta=fila[8] or '',
                fc=fila[9] or '',
                fr=fila[10] or '',
                tm=fila[11] or '',
                spo2=fila[12] or '',
                glasg=fila[13] or '15/15',
                eps=fila[14] or '',
                institucion_reporta=fila[15] or '',
                municipio=fila[16] or '',
                medico_refiere=fila[17] or '',
                medico_hudn=fila[18] or '',
                radio_operador=fila[19] or '',
                observacion=fila[20] or '',
                aceptado=fila[21] or 'NO',
                fecha_res=fila[22],
                created_by=usuario,
            )

    return {'ok': True, 'importados': len(filas)}


# ---------------------------------------------------------------------------
# 3.6 — Exportación a Excel
# ---------------------------------------------------------------------------
from io import BytesIO
from django.utils.timezone import is_aware
import openpyxl


def _excel_dt(dt):
    """
    Convierte datetimes timezone-aware en datetimes compatibles con Excel.
    Excel no soporta tzinfo.
    """
    if dt is None:
        return None

    return dt.replace(tzinfo=None) if is_aware(dt) else dt


def exportar_a_excel(queryset) -> BytesIO:
    """
    Genera un archivo Excel con todos los campos de Remision,
    incluyendo la columna OPORTUNIDAD calculada dinámicamente.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Remisiones'

    encabezados = COLUMNAS_ESPERADAS + ['OPORTUNIDAD']
    ws.append(encabezados)

    for r in queryset:
        oportunidad = calcular_oportunidad(r.fecha, r.fecha_res) if r.fecha_res else ''

        ws.append([
            _excel_dt(r.fecha), r.nombre, r.tipo_doc, r.doc, r.sexo,
            r.especialidad, r.edad, r.gest, r.diagnostico, r.ta, r.fc, r.fr, r.tm, r.spo2, r.glasg,
            r.eps, r.institucion_reporta, r.municipio, r.medico_refiere, r.medico_hudn,
            r.radio_operador, r.observacion, r.aceptado, _excel_dt(r.fecha_res),
            oportunidad,
        ])

    # Formato visual para columnas fecha y fecha_res
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].number_format = 'DD/MM/YYYY HH:MM'   # fecha
        row[23].number_format = 'DD/MM/YYYY HH:MM'  # fecha_res

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


# ---------------------------------------------------------------------------
# 3.7 — Utilidades de usuario
# ---------------------------------------------------------------------------

def generar_password_temporal() -> str:
    """
    Genera una contraseña temporal de 12 caracteres usando letras y dígitos.
    Usa el módulo `secrets` para garantizar aleatoriedad criptográficamente segura.
    """
    alfabeto = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alfabeto) for _ in range(12))


def enviar_email_recuperacion(usuario, password_temp: str) -> None:
    """
    Envía un correo electrónico al usuario con su contraseña temporal.
    Usa el backend de email configurado en Django (SMTP en producción,
    locmem en pruebas).
    """
    asunto = 'Recuperación de contraseña — CRUE Remisiones'
    mensaje = (
        f'Hola {usuario.get_full_name() or usuario.username},\n\n'
        f'Tu contraseña temporal es: {password_temp}\n\n'
        'Por favor, inicia sesión y cámbiala de inmediato.\n\n'
        'Sistema CRUE Remisiones Pacientes'
    )
    send_mail(
        subject=asunto,
        message=mensaje,
        from_email=None,  # Usa DEFAULT_FROM_EMAIL de settings
        recipient_list=[usuario.email],
        fail_silently=False,
    )
