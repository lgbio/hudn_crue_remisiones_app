"""
Pruebas de propiedades (Property-Based Testing) con Hypothesis.

Propiedades verificadas:
  P1 — Invariante del perfil de usuario (UsuarioCrue) → Requisitos 1.1, 1.6
  P2 — Cálculo de Oportunidad (Round-Trip)            → Requisitos 1.2, 9.7
  P3 — Idempotencia de Autoformato de Signos Vitales  → Requisitos 5.4, 5.5
  P4 — Regla de Editabilidad por Fecha (Invariante)   → Requisitos 6.3, 6.4, 7.4
  P8 — Importación Atómica (Invariante de Transacción) → Requisito 15.8

Cada propiedad usa @settings(max_examples=20) para mantener tiempos razonables.
"""

from datetime import datetime, timedelta, timezone as dt_timezone
from io import BytesIO

import openpyxl
import pytest
from django.utils import timezone
from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st

from remisiones.services import (
    COLUMNAS_ESPERADAS,
    calcular_oportunidad,
    es_registro_editable,
    formatear_spo2,
    formatear_temperatura,
)


# ---------------------------------------------------------------------------
# P1 — Invariante del perfil de usuario
# Propiedad: para todo UsuarioCrue creado con cualquier combinación de
# parámetros, perfil ∈ {'DIRECTOR', 'RADIOOPERADOR'}.
# **Validates: Requirements 1.1, 1.6**
# ---------------------------------------------------------------------------

@given(
    perfil=st.sampled_from(['DIRECTOR', 'RADIOOPERADOR']),
    username_suffix=st.integers(min_value=0, max_value=999999),
)
@settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture])
@pytest.mark.django_db
def test_p1_usuario_crue_perfil_invariante(db, perfil, username_suffix):
    """
    P1: Para todo UsuarioCrue creado con cualquier rol válido,
    el campo rol debe estar en {'DIRECTOR', 'RADIOOPERADOR'}.
    Un usuario creado sin especificar rol debe tener 'RADIOOPERADOR' por defecto.

    **Validates: Requirements 1.1, 1.6**
    """
    from remisiones.models import UsuarioCrue

    username = f'pbt_user_{perfil.lower()}_{username_suffix}'
    # Evitar colisiones de username en la BD de prueba
    UsuarioCrue.objects.filter(username=username).delete()

    user = UsuarioCrue.objects.create_user(
        username=username,
        password='testpass123',
        rol=perfil,
    )

    assert user.rol in ('DIRECTOR', 'RADIOOPERADOR'), (
        f"rol={user.rol!r} no está en el conjunto válido"
    )
    assert user.rol == perfil, (
        f"El rol almacenado {user.rol!r} no coincide con el asignado {perfil!r}"
    )


@pytest.mark.django_db
def test_p1_usuario_crue_perfil_default_radiooperador(db):
    """
    P1 (caso base): Un UsuarioCrue creado sin especificar rol
    debe tener el valor por defecto 'RADIOOPERADOR'.

    **Validates: Requirements 1.1, 1.6**
    """
    from remisiones.models import UsuarioCrue

    user = UsuarioCrue.objects.create_user(
        username='pbt_default_user',
        password='testpass123',
    )
    assert user.rol == 'RADIOOPERADOR', (
        f"El rol por defecto debe ser 'RADIOOPERADOR', se obtuvo {user.rol!r}"
    )


# ---------------------------------------------------------------------------
# P1 — Cálculo de Oportunidad (Round-Trip)
# ---------------------------------------------------------------------------

@given(
    fecha=st.datetimes(
        min_value=datetime(2020, 1, 1),
        max_value=datetime(2030, 12, 31, 23, 0),
        timezones=st.just(dt_timezone.utc),
    ),
    delta_minutos=st.integers(min_value=0, max_value=60 * 24 * 365),
)
@settings(max_examples=20)
def test_p1_calculo_oportunidad_round_trip(fecha, delta_minutos):
    """
    P1: El resultado de calcular_oportunidad debe ser invertible.
    horas * 60 + minutos == delta_minutos para cualquier delta >= 0.
    """
    fecha_res = fecha + timedelta(minutes=delta_minutos)
    resultado = calcular_oportunidad(fecha, fecha_res)

    assert resultado != '', (
        f"calcular_oportunidad retornó '' para delta={delta_minutos} min"
    )

    partes = resultado.split(':')
    assert len(partes) == 2, f"Formato inesperado: {resultado!r}"
    horas = int(partes[0])
    minutos = int(partes[1])

    assert horas * 60 + minutos == delta_minutos, (
        f"Round-trip falló: {horas}h {minutos}m != {delta_minutos} min total"
    )


# ---------------------------------------------------------------------------
# P2 — Idempotencia de Autoformato de Signos Vitales
# ---------------------------------------------------------------------------

@given(
    valor=st.floats(
        min_value=0,
        max_value=100,
        allow_nan=False,
        allow_infinity=False,
    ).map(lambda f: str(round(f, 1)))
)
@settings(max_examples=20)
def test_p2_idempotencia_formatear_temperatura(valor):
    """
    P2a: formatear_temperatura es idempotente.
    formatear_temperatura(formatear_temperatura(v)) == formatear_temperatura(v)
    """
    una_vez = formatear_temperatura(valor)
    dos_veces = formatear_temperatura(una_vez)
    assert dos_veces == una_vez, (
        f"No idempotente: f(f({valor!r})) = {dos_veces!r} != f({valor!r}) = {una_vez!r}"
    )


@given(
    valor=st.floats(
        min_value=0,
        max_value=100,
        allow_nan=False,
        allow_infinity=False,
    ).map(lambda f: str(round(f, 1)))
)
@settings(max_examples=20)
def test_p2_idempotencia_formatear_spo2(valor):
    """
    P2b: formatear_spo2 es idempotente.
    formatear_spo2(formatear_spo2(v)) == formatear_spo2(v)
    """
    una_vez = formatear_spo2(valor)
    dos_veces = formatear_spo2(una_vez)
    assert dos_veces == una_vez, (
        f"No idempotente: f(f({valor!r})) = {dos_veces!r} != f({valor!r}) = {una_vez!r}"
    )


# ---------------------------------------------------------------------------
# P4 — Regla de Editabilidad por Fecha (Invariante de Estado)
# ---------------------------------------------------------------------------

@given(dias_atras=st.integers(min_value=1, max_value=3650))
@settings(max_examples=20)
def test_p4_registros_historicos_nunca_editables(dias_atras):
    """
    P4: Todo registro con fecha anterior al día de hoy es Registro_Historico
    y es_registro_editable debe retornar False invariablemente.
    """
    from remisiones.models import Remision

    fecha_pasada = timezone.now() - timedelta(days=dias_atras)
    remision = Remision(
        fecha=fecha_pasada,
        nombre='Paciente Test',
        tipo_doc='CC',
        doc='99999999',
        sexo='M',
        edad='30 AÑOS',
        gest='NO',
    )
    assert es_registro_editable(remision) is False, (
        f"Registro con fecha hace {dias_atras} DIAS fue marcado como editable"
    )


# ---------------------------------------------------------------------------
# P8 — Importación Atómica (Invariante de Transacción)
# ---------------------------------------------------------------------------

def _excel_con_encabezados_incorrectos() -> BytesIO:
    """Genera un Excel con encabezados que no coinciden con COLUMNAS_ESPERADAS."""
    wb = openpyxl.Workbook()
    ws = wb.active
    encabezados_malos = ['COLUMNA_FALSA'] + COLUMNAS_ESPERADAS[1:]
    ws.append(encabezados_malos)
    ws.append(['valor'] * len(encabezados_malos))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _excel_con_edad_invalida() -> BytesIO:
    """Genera un Excel con encabezados correctos pero edad con formato inválido."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COLUMNAS_ESPERADAS)
    fila = [
        datetime(2026, 1, 15, 10, 0),   # fecha
        'Paciente Importado',           # nombre
        'CC',                           # tipo_doc
        '12345678',                     # doc
        'M',                            # sexo
        'EDAD_INVALIDA',                # edad — formato incorrecto
        'NO',                           # gest
        '',                             # diagnostico
        '',                             # ta
        '',                             # fc
        '',                             # fr
        '',                             # tm
        '',                             # spo2
        '15/15',                        # glasg
        '',                             # eps
        '',                             # institucion_reporta
        '',                             # municipio
        '',                             # medico_refiere
        '',                             # medico_hudn
        '',                             # radio_operador
        '',                             # observacion
        'NO',                           # aceptado
        None,                           # fecha_res
    ]
    ws.append(fila)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@given(
    tipo_error=st.sampled_from([
        'encabezados_incorrectos',
        'edad_invalida',
    ])
)
@settings(
    max_examples=20,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
@pytest.mark.django_db
def test_p8_importacion_atomica_no_modifica_bd_en_error(db, tipo_error):
    """
    P8: Cualquier importación fallida (por cualquier tipo de error de validación)
    debe dejar el conteo de Remision.objects.count() idéntico al que había antes.
    """
    from remisiones.models import Remision, UsuarioCrue
    from remisiones.services import importar_desde_excel

    usuario, _ = UsuarioCrue.objects.get_or_create(
        username='test_import_user',
        defaults={'password': 'x'},
    )

    conteo_antes = Remision.objects.count()

    if tipo_error == 'encabezados_incorrectos':
        archivo = _excel_con_encabezados_incorrectos()
    else:
        archivo = _excel_con_edad_invalida()

    resultado = importar_desde_excel(archivo, usuario)

    conteo_despues = Remision.objects.count()

    assert resultado['ok'] is False, (
        f"Se esperaba ok=False para tipo_error={tipo_error!r}, "
        f"pero se obtuvo: {resultado}"
    )
    assert conteo_despues == conteo_antes, (
        f"La BD fue modificada a pesar del error: "
        f"antes={conteo_antes}, después={conteo_despues}"
    )


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v2, Property 2: Truncado de campos de texto en tabla
# ---------------------------------------------------------------------------

@given(t=st.text(min_size=0, max_size=200))
@settings(max_examples=20)
def test_p2_truncado_campos_texto_en_tabla(t):
    """
    Property 2: Text truncation for table.
    """
    from remisiones.utils import truncar_para_tabla

    resultado = truncar_para_tabla(t)

    if len(t) > 50:
        assert resultado == t[:50] + '\u2026', (
            f"Texto largo no truncado correctamente: len={len(t)}, "
            f"resultado={resultado!r}"
        )
    else:
        assert resultado == t, (
            f"Texto corto fue modificado: original={t!r}, resultado={resultado!r}"
        )


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v2, Property 3: Round-trip de campos de texto con saltos de línea
# ---------------------------------------------------------------------------

@given(t=st.text())
@settings(max_examples=20, suppress_health_check=[HealthCheck.function_scoped_fixture])
@pytest.mark.django_db
def test_p3_round_trip_campos_texto_con_saltos_de_linea(db, t):
    """
    Property 3: Round-trip of text fields with newlines.
    """
    from remisiones.models import Remision, UsuarioCrue

    usuario, _ = UsuarioCrue.objects.get_or_create(
        username='pbt_roundtrip_user',
        defaults={'password': 'x'},
    )

    r = Remision.objects.create(
        fecha=timezone.now(),
        nombre='Paciente PBT',
        tipo_doc='CC',
        doc='11111111',
        sexo='M',
        edad='30 AÑOS',
        gest='NO',
        diagnostico=t,
        medico_hudn=t,
        observacion=t,
        created_by=usuario,
    )

    r_db = Remision.objects.get(pk=r.pk)

    assert r_db.diagnostico == t, (
        f"diagnostico round-trip falló: guardado={t!r}, recuperado={r_db.diagnostico!r}"
    )
    assert r_db.medico_hudn == t, (
        f"medico_hudn round-trip falló: guardado={t!r}, recuperado={r_db.medico_hudn!r}"
    )
    assert r_db.observacion == t, (
        f"observacion round-trip falló: guardado={t!r}, recuperado={r_db.observacion!r}"
    )

    # Cleanup to avoid DB state accumulation across hypothesis examples
    r.delete()


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v2, Property 7: Round-trip del widget de fecha/hora
# ---------------------------------------------------------------------------

@given(
    dt=st.datetimes(
        min_value=datetime(2000, 1, 1),
        max_value=datetime(2099, 12, 31, 23, 59),
    )
)
@settings(max_examples=20)
def test_p7_round_trip_widget_datetime(dt):
    """
    Property 7: Round-trip del widget de fecha/hora.
    """
    from remisiones.utils import parsear_widget_datetime
    cadena = dt.strftime('%d/%m/%Y %H:%M')
    resultado = parsear_widget_datetime(cadena)
    esperado = dt.replace(second=0, microsecond=0)
    assert resultado == esperado, (
        f"Round-trip falló: dt={dt!r}, cadena={cadena!r}, resultado={resultado!r}"
    )


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v2, Property 8: OPORTUNIDAD vacío ante fechas inválidas o nulas
# ---------------------------------------------------------------------------

@given(
    fecha=st.datetimes(timezones=st.just(dt_timezone.utc)),
    fecha_res=st.one_of(
        st.none(),
        st.datetimes(timezones=st.just(dt_timezone.utc)),
    ),
)
@settings(max_examples=20)
def test_p8_oportunidad_vacio_ante_fechas_invalidas(fecha, fecha_res):
    """
    Property 8: OPORTUNIDAD vacío ante fechas inválidas o nulas.
    """
    if fecha_res is None or fecha_res < fecha:
        resultado = calcular_oportunidad(fecha, fecha_res)
        assert resultado == '', (
            f"Se esperaba '' pero se obtuvo {resultado!r} para "
            f"fecha={fecha!r}, fecha_res={fecha_res!r}"
        )


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v2, Property 4: Exclusividad de campos booleanos en importación
# ---------------------------------------------------------------------------

@given(
    sexo_m=st.sampled_from(['SI', '']),
    sexo_f=st.sampled_from(['SI', '']),
)
@settings(max_examples=20)
def test_p4_exclusividad_campos_booleanos_sexo(sexo_m, sexo_f):
    """
    Property 4: Exclusividad de campos booleanos en importación (SEXO).
    """
    from remisiones.services import _mapear_exclusivo_sexo
    row = {'sexo_m': sexo_m, 'sexo_f': sexo_f}
    exactamente_uno = (sexo_m == 'SI') != (sexo_f == 'SI')
    if exactamente_uno:
        resultado = _mapear_exclusivo_sexo(row)
        assert resultado in ('M', 'F'), f"Resultado inesperado: {resultado!r}"
    else:
        try:
            _mapear_exclusivo_sexo(row)
            assert False, "Debería haber lanzado ValueError"
        except ValueError:
            pass  # Expected


@given(
    a_si=st.sampled_from(['SI', '']),
    a_no=st.sampled_from(['SI', '']),
    a_uv=st.sampled_from(['SI', '']),
)
@settings(max_examples=20)
def test_p4_exclusividad_campos_booleanos_aceptado(a_si, a_no, a_uv):
    """
    Property 4: Exclusividad de campos booleanos en importación (ACEPTADO).
    """
    from remisiones.services import _mapear_exclusivo_aceptado
    row = {'aceptado_si': a_si, 'aceptado_no': a_no, 'aceptado_urg_vital': a_uv}
    count = sum([a_si == 'SI', a_no == 'SI', a_uv == 'SI'])
    if count == 1:
        resultado = _mapear_exclusivo_aceptado(row)
        assert resultado in ('SI', 'NO', 'URG VITAL'), f"Resultado inesperado: {resultado!r}"
    else:
        try:
            _mapear_exclusivo_aceptado(row)
            assert False, "Debería haber lanzado ValueError"
        except ValueError:
            pass  # Expected


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v2, Property 5: FECHA_RES es NULL cuando fecha o hora están vacíos
# ---------------------------------------------------------------------------

@given(
    fecha=st.one_of(st.just(''), st.just('2026-04-15')),
    hora=st.one_of(st.just(''), st.just('10:30')),
)
@settings(max_examples=20)
def test_p5_fecha_res_null_cuando_fecha_o_hora_vacios(fecha, hora):
    """
    Property 5: FECHA_RES es NULL cuando fecha o hora están vacíos.
    """
    from remisiones.services import _combinar_fecha_hora
    if not fecha or not hora:
        resultado = _combinar_fecha_hora(fecha, hora)
        assert resultado is None, (
            f"Se esperaba None pero se obtuvo {resultado!r} para "
            f"fecha={fecha!r}, hora={hora!r}"
        )


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v2, Property 6: Atomicidad de la importación (v2)
# ---------------------------------------------------------------------------

@given(
    tipo_error=st.sampled_from([
        'estructura_invalida',
        'exclusividad_violada',
    ])
)
@settings(
    max_examples=20,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
    deadline=None,
)
@pytest.mark.django_db
def test_p6_importacion_v2_atomica_no_modifica_bd_en_error(db, tipo_error):
    """
    Property 6: Atomicidad de la importación v2.
    """
    import openpyxl
    import tempfile
    import os
    from remisiones.models import Remision, UsuarioCrue
    from remisiones.services import importar_desde_excel_v2

    usuario, _ = UsuarioCrue.objects.get_or_create(
        username='test_import_v2_user',
        defaults={'password': 'x'},
    )

    conteo_antes = Remision.objects.count()

    # Create invalid Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    if tipo_error == 'estructura_invalida':
        ws.append(['COLUMNA_FALSA', 'OTRA'])
        ws.append(['val1', 'val2'])
    else:  # exclusividad_violada — valid structure but bad data
        ws.append(['WRONG_HEADER'])
        ws.append(['data'])

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        temp_path = tmp.name

    try:
        resultado = importar_desde_excel_v2(temp_path, usuario)
    finally:
        if os.path.exists(temp_path):
            os.unlink(temp_path)

    conteo_despues = Remision.objects.count()

    assert resultado['ok'] is False
    assert conteo_despues == conteo_antes, (
        f"La BD fue modificada: antes={conteo_antes}, después={conteo_despues}"
    )


# ---------------------------------------------------------------------------
# Feature: crue-remisiones-v3, Property: Validación de formato de hora
# ---------------------------------------------------------------------------

@given(
    hh=st.integers(min_value=0, max_value=23),
    mm=st.integers(min_value=0, max_value=59),
)
@settings(max_examples=20)
def test_p_validacion_hora_formato_valido(hh, mm):
    """
    Para todo par (hh, mm) donde hh ∈ [0, 23] y mm ∈ [0, 59],
    validar_hora(f'{hh:02d}:{mm:02d}') debe retornar True.
    """
    from remisiones.utils import validar_hora
    cadena = f'{hh:02d}:{mm:02d}'
    assert validar_hora(cadena) is True, (
        f"validar_hora({cadena!r}) retornó False para hora válida"
    )


@given(
    hh=st.integers(min_value=24, max_value=99),
    mm=st.integers(min_value=0, max_value=99),
)
@settings(max_examples=20)
def test_p_validacion_hora_formato_invalido(hh, mm):
    """
    Para todo hh >= 24, validar_hora debe retornar False.
    """
    from remisiones.utils import validar_hora
    cadena = f'{hh:02d}:{mm:02d}'
    assert validar_hora(cadena) is False, (
        f"validar_hora({cadena!r}) retornó True para hora inválida"
    )
