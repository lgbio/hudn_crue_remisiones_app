#!/usr/bin/env python
"""
excel_to_csv_fralg062.py
========================
Validates the structure of the FRALG-062 Excel sheet and exports
only the data rows to a clean CSV file.

Expected sheet structure
------------------------
Row  1  : Title block (col D: "REMISIONES DE USUARIOS COMENTADOS POR CRUE, ESE, SAS, EPS")
Row  7  : Year/month label (col A: "AÑO:", col L: "MES: ")
Row  9  : Main column headers  (31 non-None columns across A–AE)
Row  10 : Sub-column headers   (17 sub-headers)
Rows 11+ : Data rows (col A is a consecutive integer row number)

Column layout (1-based Excel columns → CSV header)
---------------------------------------------------
Col  1  A   no
Col  2  B   fecha
Col  3  C   hora
Col  4  D   nombre_paciente
Col  5  E   tipo_documento
Col  6  F   identificacion
Col  7  G   gestante_si          ← sub-header "SI"  under "GESTANTE"
Col  8  H   gestante_no          ← sub-header "NO"
Col  9  I   sexo_m               ← sub-header "M"   under "SEXO EDAD"
Col 10  J   sexo_f               ← sub-header "F"
Col 11  K   edad                 ← sub-header "ED"
Col 12  L   diagnostico
Col 13  M   sv_ta                ← sub-header "TA"  under "SIGNOS VITALES"
Col 14  N   sv_fc                ← sub-header "FC"
Col 15  O   sv_fr                ← sub-header "FR"
Col 16  P   sv_temperatura       ← sub-header "T°"
Col 17  Q   sv_spo2              ← sub-header "SPO2"
Col 18  R   sv_glasgow           ← sub-header "Glasgow"
Col 19  S   eps
Col 20  T   institucion_reporta
Col 21  U   municipio
Col 22  V   medico_refiere
Col 23  W   medico_hudn_confirma
Col 24  X   radio_operador
Col 25  Y   observacion
Col 26  Z   aceptado_si          ← sub-header "SI"   under "ACEPTADO"
Col 27  AA  aceptado_no          ← sub-header "NO"
Col 28  AB  aceptado_urg_vital   ← sub-header "URG VITAL"
Col 29  AC  aceptado_fecha       ← sub-header "FECHA"
Col 30  AD  aceptado_hora        ← sub-header "HORA"
Col 31  AE  oportunidad
"""

import csv
import datetime
import re
from pathlib import Path

import openpyxl


# ── Expected structure fingerprints ───────────────────────────────────────────

# (0-based col index, expected value)  — used to validate rows 9 and 10
EXPECTED_MAIN_HEADERS: list[tuple[int, str]] = [
    (0,  "No."),
    (1,  "FECHA"),
    (2,  "HORA"),
    (3,  "NOMBRE PACIENTE"),
    (4,  "TIPO DOCUMETO"),          # original typo preserved
    (5,  "IDENTIFICACIÓN"),
    (6,  "GESTANTE"),
    (8,  "SEXO EDAD"),
    (11, "DIAGNÓSTICO"),
    (12, "SIGNOS VITALES"),
    (18, "EPS"),
    (19, "INSTITUCIÓN QUE REPORTA"),
    (20, "MUNICIPIO"),
    (21, "MÉDICO QUE REFIERE"),
    (22, "MÉDICO HUDN QUE CONFIRMA"),
    (23, "RADIO OPERADOR"),
    (24, "OBSERVACIÓN"),
    (25, "ACEPTADO"),
    (30, "OPORTUNIDAD"),
]

EXPECTED_SUB_HEADERS: list[tuple[int, str]] = [
    (4,  "CC,TI,CE,RC"),
    (6,  "SI"),
    (7,  "NO"),
    (8,  "M"),
    (9,  "F"),
    (10, "ED"),
    (12, "TA"),
    (13, "FC"),
    (14, "FR"),
    (15, "T°"),
    (16, "SPO2"),
    (17, "Glasgow"),
    (25, "SI"),
    (26, "NO"),
    (27, "URG VITAL"),
    (28, "FECHA"),
    (29, "HORA"),
]

EXPECTED_TITLE_FRAGMENT = "REMISIONES DE USUARIOS COMENTADOS POR CRUE"
EXPECTED_YEAR_LABEL  = "AÑO:"
EXPECTED_MONTH_LABEL = "MES:"

# CSV column names in output order (matches cols 1-31)
CSV_HEADERS = [
    "no", "fecha", "hora", "nombre_paciente", "tipo_documento", "identificacion",
    "gestante_si", "gestante_no",
    "sexo_m", "sexo_f", "edad",
    "diagnostico",
    "sv_ta", "sv_fc", "sv_fr", "sv_temperatura", "sv_spo2", "sv_glasgow",
    "eps", "institucion_reporta", "municipio",
    "medico_refiere", "medico_hudn_confirma", "radio_operador",
    "observacion",
    "aceptado_si", "aceptado_no", "aceptado_urg_vital",
    "aceptado_fecha", "aceptado_hora",
    "oportunidad",
]


# ── Structure validation ───────────────────────────────────────────────────────

class SheetStructureError(ValueError):
    """Raised when the sheet does not match the expected FRALG-062 layout."""


def _cell_str(val) -> str:
    """Normalise a cell value to stripped string for comparison."""
    if val is None:
        return ""
    return str(val).strip()


def _validate_structure(ws) -> None:
    """
    Inspect the worksheet and raise SheetStructureError with a clear message
    if any expected structural element is missing or wrong.
    """
    errors: list[str] = []

    def get_row(row_num: int) -> tuple:
        return next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))

    # 1. Title row (row 1, col D = index 3)
    row1 = get_row(1)
    title_val = _cell_str(row1[3]) if len(row1) > 3 else ""
    if EXPECTED_TITLE_FRAGMENT not in title_val:
        errors.append(
            f"Row 1 col D: expected title containing '{EXPECTED_TITLE_FRAGMENT}', "
            f"got '{title_val[:80]}'"
        )

    # 2. Year/month labels (row 7, col A = index 0, col L = index 11)
    row7 = get_row(7)
    if _cell_str(row7[0] if len(row7) > 0 else None) != EXPECTED_YEAR_LABEL:
        errors.append(
            f"Row 7 col A: expected '{EXPECTED_YEAR_LABEL}', "
            f"got '{_cell_str(row7[0] if row7 else None)}'"
        )
    if _cell_str(row7[11] if len(row7) > 11 else None) != EXPECTED_MONTH_LABEL:
        errors.append(
            f"Row 7 col L: expected '{EXPECTED_MONTH_LABEL}', "
            f"got '{_cell_str(row7[11] if len(row7) > 11 else None)}'"
        )

    # 3. Main column headers (row 9)
    row9 = get_row(9)
    for col_idx, expected in EXPECTED_MAIN_HEADERS:
        actual = _cell_str(row9[col_idx] if col_idx < len(row9) else None)
        if actual != expected:
            errors.append(
                f"Row 9 col {col_idx+1}: expected '{expected}', got '{actual}'"
            )

    # 4. Sub-column headers (row 10)
    row10 = get_row(10)
    for col_idx, expected in EXPECTED_SUB_HEADERS:
        actual = _cell_str(row10[col_idx] if col_idx < len(row10) else None)
        if actual != expected:
            errors.append(
                f"Row 10 col {col_idx+1}: expected sub-header '{expected}', got '{actual}'"
            )

    if errors:
        bullet_list = "\n  • ".join(errors)
        raise SheetStructureError(
            f"Sheet structure does not match FRALG-062 layout "
            f"({len(errors)} issue(s)):\n  • {bullet_list}"
        )


# ── Cell value converters ──────────────────────────────────────────────────────

def _fmt_date(val) -> str:
    if isinstance(val, datetime.datetime):
        return val.date().isoformat()
    if isinstance(val, datetime.date):
        return val.isoformat()
    return _cell_str(val)


def _fmt_time(val) -> str:
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M")
    return _cell_str(val)


def _fmt_bool(val) -> str:
    """'X'/'x'/True/1 → 'SI'  |  None → ''  |  else → 'NO'"""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "SI" if val else "NO"
    if isinstance(val, str) and val.strip().upper() == "X":
        return "SI"
    if isinstance(val, (int, float)) and val:
        return "SI"
    return ""


def _fmt_val(val) -> str:
    """Generic formatter for numeric/text cells."""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))          # avoid "68.0" for integers stored as float
    return str(val).strip()


def _fmt_oportunidad(val) -> str:
    """
    Extract only the HH:MM time component from the 'oportunidad' column.

    The column stores elapsed time (days/hours:minutes). openpyxl returns it
    in three distinct forms depending on the cell value and whether Excel
    evaluated the formula:

    1. datetime.time  (most common, ~4 030 rows)
       openpyxl parses values < 1 day directly as time objects.
       e.g. datetime.time(3, 33)  →  "03:33"

    2. datetime.datetime  (~34 rows)
       Values ≥ 1 day overflow into a datetime; openpyxl anchors them to
       Excel's epoch (1899-12-29 or 1900-01-01). The date part is irrelevant —
       only the time component carries the HH:MM information.
       e.g. datetime.datetime(1899, 12, 29, 1, 23)  →  "01:23"

    3. str  (~4 rows)
       Unevaluated formula results land as raw strings in the pattern "D/H:MM"
       or "D/HH:MM" (days / hours : minutes). We discard the day prefix and
       parse only the part after '/'. '#VALUE!' and other error strings → "".
       e.g. "0/0:07"  →  "00:07"
            "0/0:7"   →  "00:07"
            "#VALUE!"  →  ""
    """
    if val is None:
        return ""
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M")
    if isinstance(val, str):
        match = re.search(r"/(\d{1,2}):(\d{1,2})$", val.strip())
        if match:
            return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}"
        return ""   # '#VALUE!' or any unrecognised pattern
    return ""


def _row_to_csv_record(row: tuple) -> list[str]:
    """
    Map a tuple of up to 105 cell values (0-based) to a flat list of 31 CSV strings.
    Only columns 0–30 (Excel A–AE) are used.
    """
    c = list(row) + [None] * max(0, 31 - len(row))   # safety-pad to 31 elements

    return [
        _fmt_val(c[0]),           # no
        _fmt_date(c[1]),          # fecha
        _fmt_time(c[2]),          # hora
        _cell_str(c[3]),          # nombre_paciente
        _cell_str(c[4]),          # tipo_documento
        _fmt_val(c[5]),           # identificacion
        _fmt_bool(c[6]),          # gestante_si
        _fmt_bool(c[7]),          # gestante_no
        _fmt_bool(c[8]),          # sexo_m
        _fmt_bool(c[9]),          # sexo_f
        _fmt_val(c[10]),          # edad
        _cell_str(c[11]),         # diagnostico
        _cell_str(c[12]),         # sv_ta
        _fmt_val(c[13]),          # sv_fc
        _fmt_val(c[14]),          # sv_fr
        _fmt_val(c[15]),          # sv_temperatura
        _fmt_val(c[16]),          # sv_spo2
        _cell_str(c[17]),         # sv_glasgow
        _cell_str(c[18]),         # eps
        _cell_str(c[19]),         # institucion_reporta
        _cell_str(c[20]),         # municipio
        _cell_str(c[21]),         # medico_refiere
        _cell_str(c[22]),         # medico_hudn_confirma
        _cell_str(c[23]),         # radio_operador
        _cell_str(c[24]),         # observacion
        _fmt_bool(c[25]),         # aceptado_si
        _fmt_bool(c[26]),         # aceptado_no
        _fmt_bool(c[27]),         # aceptado_urg_vital
        _fmt_date(c[28]),         # aceptado_fecha
        _fmt_time(c[29]),         # aceptado_hora
        _fmt_oportunidad(c[30]),  # oportunidad  → HH:MM
    ]


# ── Main function ─────────────────────────────────────────────────────────────

def excelToDataframe (
    excel_path: str | Path,
    header_rows: int = 10,
) -> "pd.DataFrame":
    """
    Validate the FRALG-062 Excel sheet structure and return only the data rows
    as a typed pandas DataFrame.

    Parameters
    ----------
    excel_path  : Path to the source .xlsx file.
    header_rows : Number of rows to skip before data begins (default 10).

    Returns
    -------
    pd.DataFrame with columns defined in CSV_HEADERS and the following dtypes:
        no, edad           -> Int64  (nullable integer)
        sv_fc/fr/temp/spo2 -> float64
        fecha, hora, ...   -> object / str  ("YYYY-MM-DD", "HH:MM", plain text)
        gestante/sexo/
        aceptado booleans  -> object / str  ("SI" | pd.NA)
        oportunidad        -> object / str  ("HH:MM" | pd.NA)

    Raises
    ------
    FileNotFoundError   – if excel_path does not exist
    SheetStructureError – if the sheet layout does not match the expected structure
    """
    import pandas as pd

    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # ── 1. Validate structure (read_only=False needed for merged-cell metadata)
    wb_check = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = wb_check.sheetnames[0]   # always use the first sheet
    _validate_structure(wb_check[sheet_name])
    wb_check.close()

    # ── 2. Stream data rows (read_only for memory efficiency)
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    records: list[list] = []
    rows_skipped = 0

    for row in ws.iter_rows(min_row=header_rows + 1, values_only=True):
        # Stop at the first completely empty row
        if not any(v is not None for v in row):
            break
        # Skip rows without a numeric value in col A (No.)
        if not isinstance(row[0], (int, float)):
            rows_skipped += 1
            continue
        records.append(_row_to_csv_record(row))

    wb.close()

    # ── 3. Build DataFrame with proper dtypes
    df = pd.DataFrame(records, columns=CSV_HEADERS)

    # Integer columns: use Int64 (nullable) so empty strings become pd.NA
    for col in ("no", "edad"):
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

    # Float columns
    for col in ("sv_fc", "sv_fr", "sv_temperatura", "sv_spo2"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Replace empty strings with pd.NA in all remaining object columns
    #str_cols = df.select_dtypes(include=["object", "str"]).columns
    #df[str_cols] = df[str_cols].replace("", pd.NA)
    outfile = str(excel_path).split (".")[0] + ".csv"
    df.to_csv (outfile, index=False)

    return df


# ── CLI / quick test ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    excel = sys.argv[1] if len(sys.argv) > 1 else (
        "remisiones.xlsx"
    )

    try:
        df = excelToDataframe (excel)
        print(f"\u2713 {len(df)} filas, {len(df.columns)} columnas")
        print(df.dtypes.to_string())
        print(df.head(3).to_string())
    except SheetStructureError as e:
        print(f"\u2717 Estructura inv\u00e1lida:\n{e}", file=sys.stderr)
        import sys; sys.exit(1)


# ── Template filter: truncar_para_tabla ───────────────────────────────────────

def truncar_para_tabla(texto, max_chars=50):
    """
    Trunca un texto para mostrarlo en la tabla de registros.

    Si len(texto) > max_chars, retorna texto[:max_chars] + '…'.
    Si len(texto) <= max_chars, retorna texto sin modificar.

    Registrado como filtro de template Django en templatetags/remisiones_extras.py.
    """
    if texto is None:
        return ''
    if len(texto) > max_chars:
        return texto[:max_chars] + '\u2026'
    return texto


# ── parsear_widget_datetime ────────────────────────────────────────────────────

def parsear_widget_datetime(cadena):
    """
    Parsea una cadena en formato 'DD/MM/YYYY HH:MM' y retorna un datetime
    con segundos y microsegundos en cero.
    Retorna None si la cadena está vacía o tiene formato inválido.
    """
    from datetime import datetime as _datetime
    if not cadena or not cadena.strip():
        return None
    try:
        return _datetime.strptime(cadena.strip(), '%d/%m/%Y %H:%M')
    except ValueError:
        return None


def validar_hora(cadena: str) -> bool:
    """
    Valida formato HH:MM en 24 horas.
    HH: 00-23, MM: 00-59.
    Retorna False si la cadena está vacía o tiene formato inválido.
    """
    import re
    if not cadena:
        return False
    return bool(re.match(r'^([01]\d|2[0-3]):([0-5]\d)$', cadena.strip()))
